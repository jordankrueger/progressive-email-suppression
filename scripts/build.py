#!/usr/bin/env python3
"""Rebuild the data/ files from the local sources and upstream community lists.

Run:     python3 scripts/build.py
Options: --no-fetch   (skip upstream HTTP pulls; use local snapshots only)
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
SOURCE_XLSX = REPO / "sources" / "original-community-list.xlsx"
DATA_DIR = REPO / "data"

UPSTREAM_URLS = {
    "disposable-email-domains": (
        "https://raw.githubusercontent.com/disposable-email-domains/"
        "disposable-email-domains/master/disposable_email_blocklist.conf"
    ),
    "mailchecker": (
        "https://raw.githubusercontent.com/FGRibreau/mailchecker/master/list.txt"
    ),
    "fakefilter": (
        "https://raw.githubusercontent.com/7c/fakefilter/main/txt/data.txt"
    ),
}

DOMAIN_RE = re.compile(r"^[a-z0-9][a-z0-9.\-]*\.[a-z0-9\-]{2,}$")

TYPO_PATTERNS = [
    re.compile(p) for p in [
        r".*g?mail[a-z0-9]*\.com$",
        r".*yahoo[a-z0-9]*\.[a-z]{2,}$",
        r".*hotmail[a-z0-9]*\.[a-z]{2,}$",
        r".*icloud[a-z0-9]*\.com$",
        r".*aol[a-z0-9]*\.com$",
        r".*outlook[a-z0-9]*\.com$",
    ]
]


def normalize(raw: str) -> str | None:
    if not raw:
        return None
    d = str(raw).strip().lower()
    d = d.lstrip("@")
    if not d or d.startswith("."):
        return None
    if not DOMAIN_RE.match(d):
        return None
    return d


def read_xlsx_sheet(path: Path, sheet: str) -> set[str]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        return set()
    ws = wb[sheet]
    domains = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            d = normalize(str(cell))
            if d:
                domains.add(d)
    return domains


def fetch_upstream(url: str, timeout: int = 20) -> set[str]:
    req = urllib.request.Request(url, headers={"User-Agent": "progressive-email-suppression/0.1"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        text = r.read().decode("utf-8", errors="replace")
    out = set()
    for line in text.splitlines():
        line = line.split("#", 1)[0].strip()
        d = normalize(line)
        if d:
            out.add(d)
    return out


def classify_typo(domain: str) -> bool:
    return any(p.match(domain) for p in TYPO_PATTERNS)


def write_list(path: Path, domains: set[str], header: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    body = "\n".join(sorted(domains))
    path.write_text(f"{header}\n{body}\n", encoding="utf-8")


def build(fetch: bool = True) -> dict[str, int]:
    credo = read_xlsx_sheet(SOURCE_XLSX, "CREDO")
    avaaz = read_xlsx_sheet(SOURCE_XLSX, "Avaaz")
    all_community = read_xlsx_sheet(SOURCE_XLSX, "ALL")

    combined: set[str] = set()
    combined.update(credo)
    combined.update(avaaz)
    combined.update(all_community)

    upstream_summary = {}
    if fetch:
        for name, url in UPSTREAM_URLS.items():
            try:
                upstream = fetch_upstream(url)
                combined.update(upstream)
                upstream_summary[name] = len(upstream)
            except Exception as e:  # noqa: BLE001 — soft-fail on network errors
                print(f"[warn] skipping {name}: {e}", file=sys.stderr)
                upstream_summary[name] = -1

    typos = {d for d in combined if classify_typo(d)}

    today = dt.date.today().isoformat()
    stamp = (
        f"# progressive-email-suppression — generated {today}\n"
        f"# Source: https://github.com/jordankrueger/progressive-email-suppression\n"
        f"# License: CC0 1.0 (public domain)"
    )

    write_list(DATA_DIR / "combined.txt", combined,
               f"{stamp}\n# {len(combined)} domains (CREDO + Avaaz + upstream community lists)")
    write_list(DATA_DIR / "credo.txt", credo,
               f"{stamp}\n# CREDO Action internal list, circa 2019 ({len(credo)} domains)")
    write_list(DATA_DIR / "avaaz.txt", avaaz,
               f"{stamp}\n# Avaaz internal list ({len(avaaz)} domains)")
    write_list(DATA_DIR / "typos.txt", typos,
               f"{stamp}\n# {len(typos)} domains matching typosquat patterns of major providers")

    return {
        "combined": len(combined),
        "credo": len(credo),
        "avaaz": len(avaaz),
        "typos": len(typos),
        **{f"upstream:{k}": v for k, v in upstream_summary.items()},
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--no-fetch", action="store_true",
                        help="Skip upstream HTTP pulls")
    args = parser.parse_args()
    summary = build(fetch=not args.no_fetch)
    print("Build summary:")
    for k, v in summary.items():
        print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
